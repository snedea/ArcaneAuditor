#!/usr/bin/env python3
"""
FastAPI server for the HTML frontend of Arcane Auditor.
This serves the simple HTML/JavaScript interface with FastAPI backend.
"""

from contextlib import asynccontextmanager
import sys
import tempfile
import threading
import time
import uuid
import asyncio
from pathlib import Path
from typing import Dict, Optional
import webbrowser

# Add the project root to Python path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

# FastAPI imports
from fastapi import FastAPI, UploadFile, File, HTTPException, Form, Response
from fastapi.responses import FileResponse, HTMLResponse
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
import uvicorn

# Configuration constants
MAX_FILE_SIZE = 100 * 1024 * 1024  # 100MB default limit
CHUNK_SIZE = 8192  # 8KB chunks for streaming

# Global job management for async analysis
analysis_jobs: Dict[str, 'AnalysisJob'] = {}
job_lock = threading.Lock()

def get_dynamic_config_info():
    """Dynamically discover configuration information from all config directories."""
    config_info = {}
    
    # Search in priority order: personal, teams, presets
    config_dirs = [
        project_root / "config" / "personal",
        project_root / "config" / "teams", 
        project_root / "config" / "presets"
    ]
    
    for config_dir in config_dirs:
        if not config_dir.exists():
            continue
            
        # Search for JSON files in the directory (not subdirectories)
        for config_file in config_dir.glob("*.json"):
            config_name = config_file.stem
        
            try:
                import json
                with open(config_file, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                
                # Count enabled rules
                enabled_rules = 0
                if 'rules' in config_data:
                    for rule_name, rule_config in config_data['rules'].items():
                        # Check if rule is explicitly enabled (defaults to True if not present)
                        is_enabled = rule_config.get('enabled', True)
                        if is_enabled:
                            enabled_rules += 1
                
                # Determine performance level based on rule count
                if enabled_rules <= 10:
                    performance = "Fast"
                elif enabled_rules <= 25:
                    performance = "Balanced"
                else:
                    performance = "Thorough"
                
                # Determine config type based on directory
                if config_dir.name == "personal":
                    config_type = "Personal"
                    description = f"Personal configuration with {enabled_rules} rules enabled"
                elif config_dir.name == "teams":
                    config_type = "Team"
                    description = f"Team configuration with {enabled_rules} rules enabled"
                elif config_dir.name == "presets":
                    config_type = "Built-in"
                    description = f"Built-in configuration with {enabled_rules} rules enabled"
                
                # Create unique key by combining config name with directory type
                # This allows showing all versions (personal, team, built-in) in the UI
                config_key = f"{config_name}_{config_dir.name}"
                config_info[config_key] = {
                    "name": config_name.replace('_', ' ').title(),
                    "description": description,
                    "rules_count": enabled_rules,
                    "performance": performance,
                    "type": config_type,
                    "path": str(config_file.relative_to(project_root)),
                    "id": config_name,  # Original name for API compatibility
                    "source": config_dir.name,  # Track which directory it came from
                    "rules": config_data.get('rules', {})  # Include actual rules data
                }
                
            except Exception as e:
                print(f"Warning: Failed to load config {config_name}: {e}")
    
    return config_info

def cleanup_orphaned_files():
    """Clean up orphaned files from previous server runs."""
    uploads_dir = Path(__file__).parent / "uploads"
    if uploads_dir.exists():
        current_time = time.time()
        files_found = 0
        files_deleted = 0
        
        for file_path in uploads_dir.glob("*.zip"):
            files_found += 1
            file_age = current_time - file_path.stat().st_mtime
            # Delete files older than 1 hour
            if file_age > 3600:
                try:
                    file_path.unlink()
                    files_deleted += 1
                    print(f"Cleaned up orphaned file: {file_path.name} (age: {file_age/60:.1f} minutes)")
                except Exception as e:
                    print(f"Failed to clean up {file_path.name}: {e}")
            else:
                print(f"Keeping file: {file_path.name} (age: {file_age/60:.1f} minutes)")
        
        print(f"Cleanup summary: {files_found} files found, {files_deleted} files deleted")


class AnalysisJob:
    """Represents an analysis job with status tracking."""
    def __init__(self, job_id: str, zip_path: Path = None, config: str = "default"):
        self.job_id = job_id
        self.zip_path = zip_path
        self.config = config
        self.status = "queued"  # queued, running, completed, failed
        self.result = None
        self.error = None
        self.start_time = None
        self.end_time = None
        self.thread = None
        self.is_zip = True  # True for ZIP files, False for individual files
        self.individual_files = []  # List of Path objects for individual files
    
    def to_dict(self):
        """Convert job to dictionary for JSON serialization."""
        return {
            "job_id": self.job_id,
            "status": self.status,
            "result": self.result,
            "error": self.error,
            "start_time": self.start_time,
            "end_time": self.end_time,
            "config": self.config
        }

class JobStatusResponse(BaseModel):
    """Response model for job status."""
    job_id: str
    status: str
    result: Optional[dict] = None
    error: Optional[str] = None
    start_time: Optional[float] = None
    end_time: Optional[float] = None
    config: Optional[str] = None

class AnalysisRequest(BaseModel):
    """Request model for analysis."""
    job_id: str


@asynccontextmanager
async def lifespan(app: FastAPI):
    """Clean up orphaned files on server startup."""
    cleanup_orphaned_files()
    
    # Start periodic cleanup task
    import asyncio
    asyncio.create_task(periodic_cleanup())
    yield

# Initialize FastAPI app
app = FastAPI(
    title="Arcane Auditor API",
    description="Workday Extend Code Review Tool API",
    version="0.4.0",
    lifespan=lifespan
)

async def periodic_cleanup():
    """Run cleanup every 5 minutes."""
    while True:
        await asyncio.sleep(300)  # 5 minutes
        cleanup_orphaned_files()
        cleanup_old_jobs()

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def cleanup_old_jobs():
    """Remove completed/failed jobs older than 1 hour."""
    current_time = time.time()
    with job_lock:
        jobs_to_remove = []
        for job_id, job in analysis_jobs.items():
            if job.status in ["completed", "failed"] and job.end_time:
                if current_time - job.end_time > 3600:  # 1 hour
                    jobs_to_remove.append(job_id)
        
        for job_id in jobs_to_remove:
            job = analysis_jobs.pop(job_id)
            # Clean up any remaining temporary files
            if job.is_zip and job.zip_path and job.zip_path.exists():
                job.zip_path.unlink()
                print(f"Cleaned up remaining ZIP file: {job.zip_path.name}")
            elif not job.is_zip and job.individual_files:
                for file_path in job.individual_files:
                    if file_path.exists():
                        file_path.unlink()
                        print(f"Cleaned up remaining file: {file_path.name}")

def run_analysis_background(job: AnalysisJob):
    """Run analysis in background thread."""
    try:
        job.status = "running"
        job.start_time = time.time()
        
        # Import analysis modules here to avoid import issues
        from file_processing.processor import FileProcessor
        from parser.app_parser import ModelParser
        from parser.rules_engine import RulesEngine
        from parser.config_manager import ConfigurationManager
            
        # Process files based on job type
        file_processing_start = time.time()
        processor = FileProcessor()
        
        if job.is_zip:
            # ZIP file mode
            source_files_map = processor.process_zip_file(job.zip_path)
            
            # Delete ZIP file immediately after successful processing
            if job.zip_path and job.zip_path.exists():
                job.zip_path.unlink()
                print(f"Deleted processed ZIP file: {job.zip_path.name}")
        else:
            # Individual files mode
            source_files_map = processor.process_individual_files(job.individual_files)
            
            # Delete individual files immediately after successful processing
            for file_path in job.individual_files:
                if file_path.exists():
                    file_path.unlink()
                    print(f"Deleted processed file: {file_path.name}")
        
        file_processing_time = time.time() - file_processing_start
        
        if not source_files_map:
            job.error = "No valid source files found"
            job.status = "failed"
            job.end_time = time.time()
            return
            
        # Create project context
        parsing_start = time.time()
        parser = ModelParser()
        context = parser.parse_files(source_files_map)
        parsing_time = time.time() - parsing_start
        
        # Run analysis with specified configuration
        config_start = time.time()
        # Use project root as base path for configuration loading
        project_root = Path(__file__).parent.parent
        config_manager = ConfigurationManager(project_root)
        config = config_manager.load_config(job.config)
        rules_engine = RulesEngine(config)
        config_time = time.time() - config_start
        
        analysis_start = time.time()
        findings = rules_engine.run(context)
        analysis_time = time.time() - analysis_start
        
        # Log performance metrics
        total_time = time.time() - job.start_time
        print(f"Performance metrics for job {job.job_id}:")
        print(f"  File processing: {file_processing_time:.2f}s")
        print(f"  Parsing: {parsing_time:.2f}s")
        print(f"  Config loading: {config_time:.2f}s")
        print(f"  Analysis: {analysis_time:.2f}s")
        print(f"  Total: {total_time:.2f}s")
            
        # Convert findings to serializable format
        result = {
            "findings": [
                {
                    "rule_id": finding.rule_id,
                    "severity": finding.severity,
                    "message": finding.message,
                    "file_path": finding.file_path,
                    "line": finding.line
                    }
                    for finding in findings
            ],
            "summary": {
                "total_findings": len(findings),
                "rules_executed": len(rules_engine.rules),
                "by_severity": {
                    "action": len([f for f in findings if f.severity == "ACTION"]),
                    "advice": len([f for f in findings if f.severity == "ADVICE"])
                }
            }
        }
        
        # Add context awareness information if available
        if context.analysis_context:
            result["context"] = context.analysis_context.to_dict()
        
        job.result = result
        job.status = "completed"
        job.end_time = time.time()
        
    except Exception as e:
        job.error = str(e)
        job.status = "failed"
        job.end_time = time.time()
        print(f"Analysis failed: {e}", file=sys.stderr)


@app.post("/api/upload")
async def upload_file(
    files: list[UploadFile] = File(...), 
    config: str = Form("default")
):
    """
    Upload file(s) for analysis.
    
    Supports:
    - Single ZIP file: Complete application archive
    - Multiple individual files: .pmd, .pod, .amd, .smd, .script files
    """
    
    # Validate we have files
    if not files:
        raise HTTPException(status_code=400, detail="No files provided")
    
    # Determine if this is a ZIP upload or individual files
    is_zip_upload = len(files) == 1 and files[0].filename.lower().endswith('.zip')
    
    # Validate configuration
    config_info = get_dynamic_config_info()
    if config not in config_info:
        raise HTTPException(status_code=400, detail=f"Invalid configuration: {config}")
    
    # Get the actual config path from the selected config key
    selected_config_info = config_info[config]
    actual_config_path = selected_config_info["path"]
    
    # Generate job ID
    job_id = str(uuid.uuid4())
    
    # Save uploaded file(s)
    uploads_dir = Path(__file__).parent / "uploads"
    uploads_dir.mkdir(exist_ok=True)
    
    try:
        if is_zip_upload:
            # ZIP file mode
            file = files[0]
            
            # Validate file size
            if file.size and file.size > MAX_FILE_SIZE:
                raise HTTPException(status_code=413, detail=f"File too large. Maximum size is {MAX_FILE_SIZE // (1024*1024)}MB")
            
            zip_path = uploads_dir / f"{job_id}.zip"
            
            # Stream file to disk
            with open(zip_path, "wb") as buffer:
                while chunk := await file.read(CHUNK_SIZE):
                    buffer.write(chunk)
            
            # Create analysis job for ZIP
            job = AnalysisJob(job_id, zip_path, actual_config_path)
            job.is_zip = True
            print(f"DEBUG: Created ZIP job {job_id} with config='{actual_config_path}'")
        
        else:
            # Individual files mode
            saved_files = []
            
            for file in files:
                # Validate file extension
                valid_extensions = ['.pmd', '.pod', '.amd', '.smd', '.script']
                if not any(file.filename.lower().endswith(ext) for ext in valid_extensions):
                    raise HTTPException(status_code=400, detail=f"Invalid file type: {file.filename}. Only {', '.join(valid_extensions)} files are allowed.")
                
                # Validate file size
                if file.size and file.size > MAX_FILE_SIZE:
                    raise HTTPException(status_code=413, detail=f"File {file.filename} too large. Maximum size is {MAX_FILE_SIZE // (1024*1024)}MB")
                
                # Save file
                file_path = uploads_dir / f"{job_id}_{file.filename}"
                with open(file_path, "wb") as buffer:
                    while chunk := await file.read(CHUNK_SIZE):
                        buffer.write(chunk)
                
                saved_files.append(file_path)
            
            # Create analysis job for individual files
            job = AnalysisJob(job_id, None, actual_config_path)
            job.is_zip = False
            job.individual_files = saved_files
            print(f"DEBUG: Created individual files job {job_id} with {len(saved_files)} files, config='{actual_config_path}'")
        
        with job_lock:
            analysis_jobs[job_id] = job
        
        # Start background analysis
        job.thread = threading.Thread(target=run_analysis_background, args=(job,))
        job.thread.daemon = True
        job.thread.start()
        
        # Cleanup old jobs
        cleanup_old_jobs()
        
        return {"job_id": job_id, "status": "queued", "config": actual_config_path}
            
    except HTTPException:
        # Re-raise HTTP exceptions
        raise
    except Exception as e:
        # Clean up files if upload failed
        if is_zip_upload and 'zip_path' in locals() and zip_path.exists():
            zip_path.unlink()
        elif 'saved_files' in locals():
            for file_path in saved_files:
                if file_path.exists():
                    file_path.unlink()
        raise HTTPException(status_code=500, detail=f"Upload failed: {str(e)}")

@app.get("/api/configs")
async def get_available_configs(response: Response):
    """Get list of available configurations."""
    # Add cache-busting headers
    response.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response.headers["Pragma"] = "no-cache"
    response.headers["Expires"] = "0"
    
    config_info = get_dynamic_config_info()
    available_configs = []
    
    for config_name, info in config_info.items():
        config_data = info.copy()
        config_data["id"] = config_name
        available_configs.append(config_data)
    
    return {"configs": available_configs}

@app.get("/api/job/{job_id}", response_model=JobStatusResponse)
async def get_job_status(job_id: str):
    """Get the status of an analysis job."""
    with job_lock:
        if job_id not in analysis_jobs:
            raise HTTPException(status_code=404, detail="Job not found")
        
        job = analysis_jobs[job_id]
        return JobStatusResponse(**job.to_dict())

@app.get("/api/download/{job_id}")
async def download_excel(job_id: str):
    """Download analysis results as Excel file."""
    with job_lock:
        if job_id not in analysis_jobs:
            raise HTTPException(status_code=404, detail="Job not found")
        
        job = analysis_jobs[job_id]
        
        if job.status != "completed":
            raise HTTPException(status_code=400, detail="Analysis not completed")
        
        if not job.result:
            raise HTTPException(status_code=500, detail="No results available")
        
        try:
            # Use shared OutputFormatter logic for Excel generation
            from output.formatter import OutputFormatter
            from datetime import datetime
            import tempfile
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            from pathlib import Path
            
            # Create workbook using shared logic
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            
            # Group findings by file (using shared logic)
            findings_by_file = {}
            for finding_data in job.result["findings"]:
                file_path = finding_data["file_path"] or "Unknown"
                # Clean file path by removing job ID prefix
                import re
                clean_file_path = re.sub(r'^[a-f0-9-]+_', '', file_path)
                if clean_file_path not in findings_by_file:
                    findings_by_file[clean_file_path] = []
                findings_by_file[clean_file_path].append(finding_data)
            
            # Create summary sheet
            summary_sheet = wb.create_sheet("Summary")
            summary_sheet.append(["Analysis Summary"])
            summary_sheet.append(["Files Analyzed", len(findings_by_file)])
            summary_sheet.append(["Rules Executed", job.result["summary"]["rules_executed"]])
            summary_sheet.append(["Total Issues", len(job.result["findings"])])
            summary_sheet.append(["Action", len([f for f in job.result["findings"] if f["severity"] == "ACTION"])])
            summary_sheet.append(["Advice", len([f for f in job.result["findings"] if f["severity"] == "ADVICE"])])
            summary_sheet.append([])
            summary_sheet.append(["File", "Issues", "Action", "Advice"])
            
            # Style summary sheet
            summary_sheet['A1'].font = Font(bold=True, size=14)
            for row in range(1, 8):
                summary_sheet[f'A{row}'].font = Font(bold=True)
            
            # Create sheets for each file
            for file_path, file_findings in findings_by_file.items():
                # Clean sheet name (Excel has restrictions)
                sheet_name = Path(file_path).stem[:31]  # Excel sheet name limit
                sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
                if not sheet_name:
                    sheet_name = "Unknown"
                
                ws = wb.create_sheet(sheet_name)
                
                # Headers
                headers = ["Rule ID", "Severity", "Line", "Message"]
                ws.append(headers)
                
                # Style headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Add findings
                for finding in file_findings:
                    row = [
                        finding["rule_id"],
                        finding["severity"],
                        finding["line"],
                        finding["message"]
                    ]
                    ws.append(row)
                    
                    # Color code by severity
                    severity_colors = {
                        "ACTION": "FFCCCC",    # Light red
                        "ADVICE": "CCE5FF"     # Light blue
                    }
                    
                    if finding["severity"] in severity_colors:
                        fill = PatternFill(start_color=severity_colors[finding["severity"]], 
                                         end_color=severity_colors[finding["severity"]], 
                                         fill_type="solid")
                        for col_num in range(1, len(headers) + 1):
                            ws.cell(row=ws.max_row, column=col_num).fill = fill
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 150)  # Cap at 150
                    ws.column_dimensions[column_letter].width = adjusted_width
                
                # Update summary sheet
                action_count = len([f for f in file_findings if f["severity"] == "ACTION"])
                advice_count = len([f for f in file_findings if f["severity"] == "ADVICE"])
                summary_sheet.append([file_path, len(file_findings), action_count, advice_count])
            
            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
            wb.save(temp_file.name)
            temp_file.close()
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y-%m-%d")
            filename = f"arcane-auditor-results-{timestamp}.xlsx"
            
            return FileResponse(
                path=temp_file.name,
                filename=filename,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Convert web service findings to CLI format
            findings = []
            for finding_data in job.result["findings"]:
                # Create a simple Finding-like object for compatibility
                class WebFinding:
                    def __init__(self, data):
                        self.rule_id = data["rule_id"]
                        self.severity = data["severity"]
                        self.line = data["line"]
                        self.message = data["message"]
                        self.file_path = data["file_path"]
                
                findings.append(WebFinding(finding_data))
            
            # Group findings by file
            findings_by_file = {}
            for finding in findings:
                file_path = finding.file_path or "Unknown"
                # Clean file path by removing job ID prefix
                import re
                clean_file_path = re.sub(r'^[a-f0-9-]+_', '', file_path)
                if clean_file_path not in findings_by_file:
                    findings_by_file[clean_file_path] = []
                findings_by_file[clean_file_path].append(finding)
            
            # Create summary sheet
            summary_sheet = wb.create_sheet("Summary")
            summary_sheet.append(["Analysis Summary"])
            summary_sheet.append(["Files Analyzed", len(findings_by_file)])
            summary_sheet.append(["Rules Executed", job.result["summary"]["rules_executed"]])
            summary_sheet.append(["Total Issues", len(findings)])
            summary_sheet.append(["Action", len([f for f in findings if f.severity == "ACTION"])])
            summary_sheet.append(["Advices", len([f for f in findings if f.severity == "ADVICE"])])
            summary_sheet.append([])
            summary_sheet.append(["File", "Issues", "Action", "Advice"])
            
            # Style summary sheet
            summary_sheet['A1'].font = Font(bold=True, size=14)
            for row in range(1, 8):
                summary_sheet[f'A{row}'].font = Font(bold=True)
            
            # Add file summary data
            for file_path, file_findings in findings_by_file.items():
                action_count = len([f for f in file_findings if f.severity == "ACTION"])
                advice_count = len([f for f in file_findings if f.severity == "ADVICE"])
                summary_sheet.append([file_path, len(file_findings), action_count, advice_count])
            
            # Create sheets for each file
            for file_path, file_findings in findings_by_file.items():
                # Clean sheet name (Excel has restrictions)
                sheet_name = Path(file_path).stem[:31]  # Excel sheet name limit
                sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
                if not sheet_name:
                    sheet_name = "Unknown"
                
                ws = wb.create_sheet(sheet_name)
                
                # Headers
                headers = ["Rule ID", "Severity", "Line", "Message", "File Path"]
                ws.append(headers)
                
                # Style headers
                header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                for col_num, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                
                # Add findings
                for finding in file_findings:
                    row = [
                        finding.rule_id,
                        finding.severity,
                        finding.line,
                        finding.message,
                        finding.file_path
                    ]
                    ws.append(row)
                    
                    # Color code by severity
                    severity_colors = {
                        "ACTION": "FFCCCC",    # Light red
                        "ADVICE": "CCE5FF",     # Light blue
                    }
                    
                    if finding.severity in severity_colors:
                        fill = PatternFill(start_color=severity_colors[finding.severity], 
                                         end_color=severity_colors[finding.severity], 
                                         fill_type="solid")
                        for col_num in range(1, len(headers) + 1):
                            ws.cell(row=ws.max_row, column=col_num).fill = fill
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to temporary file
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
            wb.save(temp_file.name)
            temp_file.close()
            
            # Generate filename
            timestamp = datetime.now().strftime("%Y-%m-%d")
            filename = f"arcane-auditor-results-{timestamp}.xlsx"
            
            return FileResponse(
                path=temp_file.name,
                filename=filename,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Excel generation failed: {str(e)}")

@app.get("/", response_class=HTMLResponse)
async def serve_index():
    """Serve the main HTML page."""
    index_path = static_dir / "index.html"
    if not index_path.exists():
        raise HTTPException(status_code=404, detail="Index file not found")
    
    with open(index_path, 'r', encoding='utf-8') as f:
        content = f.read()
    return HTMLResponse(content=content)

@app.get("/api/health")
async def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "version": "0.1.3"}

# Mount static files at /static to avoid conflicts with API routes
static_dir = Path(__file__).parent / "frontend"
app.mount("/static", StaticFiles(directory=str(static_dir)), name="static")

def main():
    """Main function to run the server."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Arcane Auditor FastAPI Server")
    parser.add_argument("--port", type=int, default=8080, help="Port to run the server on")
    parser.add_argument("--host", type=str, default="127.0.0.1", help="Host to bind to")
    parser.add_argument("--open-browser", action="store_true", help="Open browser automatically")
    
    args = parser.parse_args()
    
    if args.open_browser:
        # Open browser after a short delay
        def open_browser():
            time.sleep(1)
            webbrowser.open(f"http://{args.host}:{args.port}")
        
        threading.Thread(target=open_browser, daemon=True).start()
    
    print(f"Starting Arcane Auditor FastAPI server on http://{args.host}:{args.port}")
    print("Press Ctrl+C to stop the server")
    
    uvicorn.run(
        app,
        host=args.host,
        port=args.port,
        log_level="info"
    )

if __name__ == "__main__":
    main()
