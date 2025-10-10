"""
Output formatting module for Arcane Auditor.
Provides various output formats with emojis and better formatting.
"""
from typing import List, Dict, Optional, TYPE_CHECKING
from pathlib import Path
import json
import re
from enum import Enum

from parser.rules.base import Finding

if TYPE_CHECKING:
    from parser.models import ProjectContext


class OutputFormat(Enum):
    """Available output formats."""
    CONSOLE = "console"
    JSON = "json"
    SUMMARY = "summary"
    EXCEL = "excel"


class OutputFormatter:
    """Formats analysis results for different output types."""
    
    # Emoji mappings for different severity levels
    SEVERITY_EMOJIS = {
        "ACTION": "🔴",
        "ADVICE": "🔵"
    }
    
    # Emoji mappings for different rule categories
    RULE_CATEGORY_EMOJIS = {
        "SCRIPT": "📝",
        "STRUCT": "🏗️",
        "STYLE": "🎨",
        "CUSTOM": "⚙️"
    }
    
    def __init__(self, format_type: OutputFormat = OutputFormat.CONSOLE):
        self.format_type = format_type
    
    def format_results(self, findings: List[Finding], total_files: int = 0, total_rules: int = 0, 
                      context: Optional['ProjectContext'] = None) -> str:
        """Format analysis results based on the selected format."""
        if self.format_type == OutputFormat.JSON:
            return self._format_json(findings, total_files, total_rules, context)
        elif self.format_type == OutputFormat.SUMMARY:
            return self._format_summary(findings, total_files, total_rules)
        elif self.format_type == OutputFormat.EXCEL:
            return self._format_excel(findings, total_files, total_rules, context)
        else:
            return self._format_console(findings, total_files, total_rules, context)
    
    def _format_console(self, findings: List[Finding], total_files: int, total_rules: int,
                        context: Optional['ProjectContext'] = None) -> str:
        """Format results for console output with emojis and better formatting."""
        output = []
        
        # Header with emojis
        output.append("🔮 **Arcane Auditor Analysis Results**")
        output.append("=" * 50)
        
        # Add context awareness panel if context available
        if context and context.analysis_context:
            context_panel = self._format_context_panel_console(context.analysis_context)
            if context_panel:
                output.append(context_panel)
                output.append("")
        
        # Summary statistics
        if total_files > 0 or total_rules > 0:
            output.append(f"📊 **Analysis Summary:**")
            output.append(f"   📁 Files analyzed: {total_files}")
            output.append(f"   🔍 Rules executed: {total_rules}")
            output.append(f"   ⚠️  Issues found: {len(findings)}")
            output.append("")
        
        if not findings:
            output.append("✅ **No issues found!** Your code looks great!")
            return "\n".join(output)
        
        # Group findings by file
        findings_by_file = self._group_findings_by_file(findings)
        
        output.append(f"🚨 **Found {len(findings)} issue(s):**")
        output.append("")
        
        for file_path, file_findings in findings_by_file.items():
            # File header with emoji
            output.append(f"📄 **{file_path}**")
            output.append("-" * (len(file_path) + 4))
            
            # Group findings by severity within each file
            severity_groups = self._group_findings_by_severity(file_findings)
            
            for severity, severity_findings in severity_groups.items():
                emoji = self.SEVERITY_EMOJIS.get(severity, "❓")
                output.append(f"  {emoji} **{severity}** ({len(severity_findings)} issue(s))")
                
                for finding in severity_findings:
                    # Rule category emoji
                    rule_category = finding.rule_id.split('0')[0] if '0' in finding.rule_id else "UNKNOWN"
                    category_emoji = self.RULE_CATEGORY_EMOJIS.get(rule_category, "🔧")
                    
                    # Format the finding with file path
                    file_display = finding.file_path.split('\\')[-1] if finding.file_path else "Unknown"
                    # Clean file path by removing job ID prefix
                    file_display = re.sub(r'^[a-f0-9-]+_', '', file_display)
                    output.append(f"    {category_emoji} **[{finding.rule_id}:{finding.line}]** in `{file_display}`: {finding.message}")
            
            output.append("")  # Empty line between files
        
        # Footer with helpful message
        output.append("💡 **Tip:** Use `--config` to load custom config files or `--help` for more options.")
        
        return "\n".join(output)
    
    def _format_summary(self, findings: List[Finding], total_files: int, total_rules: int) -> str:
        """Format a concise summary of results."""
        output = []
        
        # Quick stats
        action_count = len([f for f in findings if f.severity == "ACTION"])
        advice_count = len([f for f in findings if f.severity == "ADVICE"])
        
        if not findings:
            return "✅ No issues found!"
        
        output.append(f"🔍 Found {len(findings)} issue(s): {action_count} action, {advice_count} advice")
        
        # Show top 3 issues
        top_findings = findings[:3]
        for finding in top_findings:
            emoji = self.SEVERITY_EMOJIS.get(finding.severity, "❓")
            output.append(f"  {emoji} [{finding.rule_id}] {finding.message}")
        
        if len(findings) > 3:
            output.append(f"  ... and {len(findings) - 3} more")
        
        return "\n".join(output)
    
    def _format_json(self, findings: List[Finding], total_files: int, total_rules: int,
                    context: Optional['ProjectContext'] = None) -> str:
        """Format results as JSON."""
        result = {
            "summary": {
                "total_files": total_files,
                "total_rules": total_rules,
                "total_findings": len(findings),
                "findings_by_severity": {
                    "ACTION": len([f for f in findings if f.severity == "ACTION"]),
                    "ADVICE": len([f for f in findings if f.severity == "ADVICE"])
                }
            },
            "findings": [
                {
                    "rule_id": finding.rule_id,
                    "severity": finding.severity,
                    "message": finding.message,
                    "file_path": finding.file_path,
                    "line": finding.line
                }
                for finding in findings
            ]
        }
        
        # Add context information if available
        if context and context.analysis_context:
            result["context"] = context.analysis_context.to_dict()
        
        return json.dumps(result, indent=2)
    
    def _group_findings_by_file(self, findings: List[Finding]) -> Dict[str, List[Finding]]:
        """Group findings by file path."""
        groups = {}
        for finding in findings:
            file_path = finding.file_path or "Unknown"
            # Clean file path by removing job ID prefix
            clean_file_path = re.sub(r'^[a-f0-9-]+_', '', file_path)
            if clean_file_path not in groups:
                groups[clean_file_path] = []
            groups[clean_file_path].append(finding)
        return groups
    
    def _group_findings_by_severity(self, findings: List[Finding]) -> Dict[str, List[Finding]]:
        """Group findings by severity level."""
        groups = {}
        for finding in findings:
            severity = finding.severity
            if severity not in groups:
                groups[severity] = []
            groups[severity].append(finding)
        
        # Sort by severity priority
        severity_order = ["ACTION", "ADVICE"]
        return {k: groups[k] for k in severity_order if k in groups}
    
    def _format_context_panel_console(self, analysis_context) -> str:
        """
        Format the context awareness panel for console output.
        
        Informs users when analysis is partial due to missing cross-file dependencies.
        
        Args:
            analysis_context: AnalysisContext instance with file and check tracking
            
        Returns:
            Formatted string for console display, or empty string if complete
        """
        lines = []
        
        lines.append("━" * 60)
        lines.append(f"📁 Files Analyzed ({len(analysis_context.files_analyzed)})")
        for file_path in analysis_context.files_analyzed:
            # Remove job ID prefix if present (format: uuid_filename.ext)
            clean_file_name = re.sub(r'^[a-f0-9-]+_', '', file_path)
            lines.append(f"   ✓ {clean_file_name}")
        lines.append("")
        
        if analysis_context.is_complete:
            lines.append("✓  Context: Complete Analysis")
            lines.append("   All context files provided")
            lines.append("   All validation rules executed")
        else:
            lines.append("ℹ️  Context: Partial Analysis")
            missing = ", ".join(sorted(analysis_context.files_missing))
            lines.append(f"   Missing Context Files: {missing}")
            lines.append("")
            
            not_executed = analysis_context.rules_not_executed
            if not_executed:
                lines.append(f"   Rules Not Executed ({len(not_executed)}):")
                for rule in not_executed:
                    lines.append(f"   • {rule['rule']} - {rule['reason']}")
                lines.append("")
            
            partially = analysis_context.rules_partially_executed
            if partially:
                lines.append(f"   Rules Partially Executed ({len(partially)}):")
                for rule_name, details in partially.items():
                    skipped = ", ".join(details['skipped_checks'])
                    lines.append(f"   • {rule_name} - {skipped} skipped ({details['reason']})")
                lines.append("")
            
            lines.append("   💡 Provide AMD/SMD files for complete cross-file validation")
        
        lines.append("━" * 60)
        
        return "\n".join(lines)
    
    def _add_context_sheet_excel(self, workbook, analysis_context):
        """
        Add a context awareness sheet to Excel workbook.
        
        Args:
            workbook: openpyxl Workbook instance
            analysis_context: AnalysisContext instance with file and check tracking
        """
        from openpyxl.styles import Font, PatternFill, Alignment
        
        context_sheet = workbook.create_sheet("Context", 0)  # Insert as first sheet
        
        # Title
        context_sheet.append(["Analysis Context"])
        context_sheet['A1'].font = Font(bold=True, size=14)
        context_sheet.append([])
        
        # Analysis Type
        context_sheet.append(["Analysis Type", analysis_context.analysis_type])
        context_sheet['A3'].font = Font(bold=True)
        
        # Context Status
        status = "Complete" if analysis_context.is_complete else "Partial"
        context_sheet.append(["Context Status", status])
        context_sheet['A4'].font = Font(bold=True)
        
        # Apply color based on status
        status_fill = PatternFill(
            start_color="C6EFCE" if analysis_context.is_complete else "FFC7CE",
            end_color="C6EFCE" if analysis_context.is_complete else "FFC7CE",
            fill_type="solid"
        )
        context_sheet['B4'].fill = status_fill
        
        context_sheet.append([])
        
        # Files Analyzed
        context_sheet.append(["Files Analyzed"])
        context_sheet[f'A{context_sheet.max_row}'].font = Font(bold=True)
        for file_path in analysis_context.files_analyzed:
            # Remove job ID prefix if present (format: uuid_filename.ext)
            clean_file_name = re.sub(r'^[a-f0-9-]+_', '', file_path)
            context_sheet.append([clean_file_name])
        
        context_sheet.append([])
        
        # Files Present
        context_sheet.append(["File Types Included", ", ".join(sorted(analysis_context.files_present))])
        context_sheet[f'A{context_sheet.max_row}'].font = Font(bold=True)
        
        # Files Missing
        if analysis_context.files_missing:
            context_sheet.append(["File Types Not Included", ", ".join(sorted(analysis_context.files_missing))])
            context_sheet[f'A{context_sheet.max_row}'].font = Font(bold=True)
            # Highlight missing files in yellow
            missing_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            context_sheet[f'B{context_sheet.max_row}'].fill = missing_fill
        
        context_sheet.append([])
        
        # Rules Not Executed
        not_executed = analysis_context.rules_not_executed
        if not_executed:
            context_sheet.append([f"Rules Not Executed ({len(not_executed)})"])
            context_sheet[f'A{context_sheet.max_row}'].font = Font(bold=True)
            context_sheet.append(["Rule", "Reason"])
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            context_sheet[f'A{context_sheet.max_row}'].fill = header_fill
            context_sheet[f'A{context_sheet.max_row}'].font = header_font
            context_sheet[f'B{context_sheet.max_row}'].fill = header_fill
            context_sheet[f'B{context_sheet.max_row}'].font = header_font
            
            for rule in not_executed:
                context_sheet.append([rule['rule'], rule['reason']])
            
            context_sheet.append([])
        
        # Rules Partially Executed
        partially = analysis_context.rules_partially_executed
        if partially:
            context_sheet.append([f"Rules Partially Executed ({len(partially)})"])
            context_sheet[f'A{context_sheet.max_row}'].font = Font(bold=True)
            context_sheet.append(["Rule", "Skipped Checks", "Reason"])
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for col in ['A', 'B', 'C']:
                context_sheet[f'{col}{context_sheet.max_row}'].fill = header_fill
                context_sheet[f'{col}{context_sheet.max_row}'].font = header_font
            
            for rule_name, details in partially.items():
                skipped = ", ".join(details['skipped_checks'])
                context_sheet.append([rule_name, skipped, details['reason']])
            
            context_sheet.append([])
        
        # Recommendation
        if not analysis_context.is_complete:
            context_sheet.append(["Recommendation", "Provide AMD and SMD files for complete cross-file validation"])
            context_sheet[f'A{context_sheet.max_row}'].font = Font(bold=True, italic=True)
            context_sheet[f'B{context_sheet.max_row}'].font = Font(italic=True)
        
        # Auto-adjust column widths
        for column in context_sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 80)  # Cap at 80 for readability
            context_sheet.column_dimensions[column_letter].width = adjusted_width
    
    def _apply_accessible_styling(self, ws, findings: List[Finding]):
        """
        Apply WCAG-friendly, color-blind accessible styling to findings sheet.
        
        Uses solid colors per severity (warm orange for ACTION, cool blue for ADVICE),
        universal symbols (⚠ for ACTION, ℹ for ADVICE), borders, and fonts that work
        consistently across platforms without relying on color emoji support.
        """
        from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
        
        # --- WCAG-Friendly Color Palette ---
        # Simple, solid colors (no alternation needed - we'll use symbols for distinction)
        action_fill = PatternFill(start_color="FCD5B5", end_color="FCD5B5", fill_type="solid")  # warm orange
        advice_fill = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")  # cool blue
        other_fill = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")   # neutral gray
        
        # Borders - hairline bottom border for clean grid rhythm
        hairline = Side(style="hair", color="D6D6D6")  # Very light gray
        border = Border(bottom=hairline)
        
        # Fonts
        symbol_font = Font(name="Segoe UI Symbol", bold=True, size=10)  # For ⚠ / ℹ symbols
        bold_font = Font(name="Segoe UI", size=10, bold=True)
        mono_font = Font(name="Consolas", size=10)  # easier for reading code/paths
        normal_font = Font(name="Segoe UI", size=10)  # Default font
        wrap = Alignment(wrap_text=True, vertical="top", indent=1)  # Small indent for breathing room
        center = Alignment(horizontal="center", vertical="center")  # For severity column
        
        # Apply styling to data rows (skip header row)
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            severity = (row[1].value or "OTHER").strip().upper()
            
            # Remove symbols if present from previous formatting
            if "⚠" in severity or "ℹ" in severity:
                severity = severity.split()[-1]  # Get last word (ACTION or ADVICE)
            
            # Determine row fill based on severity
            if severity == "ACTION":
                row_fill = action_fill
            elif severity == "ADVICE":
                row_fill = advice_fill
            else:
                row_fill = other_fill
            
            for cell in row:
                cell.fill = row_fill
                cell.border = border  # Hairline bottom border for grid rhythm
                
                # Add distinctive visual cue: symbol + text in Severity column
                if cell.column == 2:  # Severity column (B)
                    if severity == "ACTION":
                        cell.value = "⚠ ACTION"  # warning symbol
                    elif severity == "ADVICE":
                        cell.value = "ℹ ADVICE"  # info symbol
                    else:
                        cell.value = severity
                    cell.font = symbol_font
                    cell.alignment = center
                
                # Bold Rule ID column
                elif cell.column == 1:  # Rule ID column (A)
                    cell.font = bold_font
                    cell.alignment = wrap
                # Monospace font for Message column (better for code/paths)
                elif cell.column == 4:  # Message column (D)
                    cell.font = mono_font
                    cell.alignment = wrap
                # Normal font for Line column
                elif cell.column == 3:  # Line column (C)
                    cell.font = normal_font
                    cell.alignment = wrap
                # Center alignment for Fixed column (dropdown)
                elif cell.column == 5:  # Fixed column (E)
                    cell.font = normal_font
                    cell.alignment = center
        
        # Adjust column dimensions for better readability
        ws.column_dimensions["A"].width = 50  # Rule ID (accommodate longest rule names)
        ws.column_dimensions["B"].width = 14  # Severity (with symbol)
        ws.column_dimensions["C"].width = 6   # Line
        ws.column_dimensions["D"].width = 120 # Message
        ws.column_dimensions["E"].width = 12  # Fixed status dropdown
        
        # Auto-adjust row heights based on wrapped text in Message column
        msg_col_width = ws.column_dimensions["D"].width or 120
        
        for row_num in range(2, ws.max_row + 1):
            # Get the Message cell (column D/4)
            msg_cell = ws.cell(row=row_num, column=4)
            
            if msg_cell.value:
                text = str(msg_cell.value)
                text_len = len(text)
                
                # Count explicit line breaks
                newline_count = text.count("\n")
                
                # Estimate wrapped lines based on column width
                # chars per line = width * 1.2 (accounting for character width variance)
                est_lines = max(1, text_len // int(msg_col_width * 1.2))
                
                # Combine explicit breaks with wrapping estimate
                total_lines = max(1, newline_count + est_lines)
                
                # Set row height: ~15 points per line, minimum 25 for breathing room
                ws.row_dimensions[row_num].height = max(total_lines * 15, 25)
            else:
                # Set minimum height for empty/short rows
                ws.row_dimensions[row_num].height = 25
    
    def _format_excel(self, findings: List[Finding], total_files: int, total_rules: int, 
                     context: Optional['ProjectContext'] = None) -> str:
        """Format results as Excel file with accessible, color-blind friendly styling."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return "Error: openpyxl package required for Excel export. Install with: pip install openpyxl"
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Add context awareness sheet if context available
        if context and context.analysis_context:
            self._add_context_sheet_excel(wb, context.analysis_context)
        
        # Group findings by file
        findings_by_file = self._group_findings_by_file(findings)
        
        # Create summary sheet
        summary_sheet = wb.create_sheet("Summary")
        summary_sheet.append(["Analysis Summary"])
        summary_sheet.append(["Files Analyzed", total_files])
        summary_sheet.append(["Rules Executed", total_rules])
        summary_sheet.append(["Total Issues", len(findings)])
        summary_sheet.append(["Action", len([f for f in findings if f.severity == "ACTION"])])
        summary_sheet.append(["Advice", len([f for f in findings if f.severity == "ADVICE"])])
        summary_sheet.append([])
        summary_sheet.append(["File", "# Issues", "# Actions", "# Advices"])
        
        # Style summary sheet
        summary_sheet['A1'].font = Font(bold=True, size=14)
        for row in range(1, 8):
            summary_sheet[f'A{row}'].font = Font(bold=True)
        
        # Create sheets for each file with accessible styling
        for file_path, file_findings in findings_by_file.items():
            # Clean sheet name (Excel has restrictions)
            sheet_name = Path(file_path).stem[:31]  # Excel sheet name limit
            sheet_name = "".join(c for c in sheet_name if c.isalnum() or c in (' ', '-', '_')).strip()
            if not sheet_name:
                sheet_name = "Unknown"
            
            ws = wb.create_sheet(sheet_name)
            
            # Headers (added "Fixed" column for tracking)
            headers = ["Rule ID", "Severity", "Line", "Message", "Fixed"]
            ws.append(headers)
            
            # Style headers with darker, more intentional background
            header_fill = PatternFill(start_color="22242A", end_color="22242A", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
            
            # Sort findings by severity (ACTION first) then by rule_id
            sorted_file_findings = sorted(file_findings, key=lambda f: (
                0 if f.severity == "ACTION" else 1,  # ACTION before ADVICE
                f.rule_id
            ))
            
            # Add findings with empty "Fixed" column
            for finding in sorted_file_findings:
                row = [
                    finding.rule_id,
                    finding.severity,
                    finding.line,
                    finding.message,
                    ""  # Empty Fixed column for user tracking
                ]
                ws.append(row)
            
            # Apply accessible styling to this sheet
            self._apply_accessible_styling(ws, sorted_file_findings)
            
            # Add data validation dropdown to "Fixed" column (E)
            from openpyxl.worksheet.datavalidation import DataValidation
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.styles import PatternFill
            
            # Create dropdown with tracking options
            dv = DataValidation(type="list", formula1='"To Do,Fixed,Won\'t Fix"', allow_blank=True)
            dv.error = 'Please select from the dropdown'
            dv.errorTitle = 'Invalid Entry'
            dv.prompt = 'Select status'
            dv.promptTitle = 'Fix Status'
            
            # Apply validation to all data rows in Fixed column (E)
            ws.add_data_validation(dv)
            dv.add(f"E2:E{ws.max_row}")
            
            # Add conditional formatting to color-code the Fixed column
            # Green for "Fixed"
            fixed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fixed_rule = CellIsRule(operator='equal', formula=['"Fixed"'], fill=fixed_fill)
            
            # Light gray for "Won't Fix"
            wont_fix_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            wont_fix_rule = CellIsRule(operator='equal', formula=['"Won\'t Fix"'], fill=wont_fix_fill)
            
            # Light yellow for "To Do"
            todo_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
            todo_rule = CellIsRule(operator='equal', formula=['"To Do"'], fill=todo_fill)
            
            # Apply conditional formatting to Fixed column range
            ws.conditional_formatting.add(f"E2:E{ws.max_row}", fixed_rule)
            ws.conditional_formatting.add(f"E2:E{ws.max_row}", wont_fix_rule)
            ws.conditional_formatting.add(f"E2:E{ws.max_row}", todo_rule)
            
            # Update summary sheet
            action_count = len([f for f in file_findings if f.severity == "ACTION"])
            advice_count = len([f for f in file_findings if f.severity == "ADVICE"])
            summary_sheet.append([file_path, len(file_findings), action_count, advice_count])
        
        # Save to temporary file and return path
        import tempfile
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        return temp_file.name
